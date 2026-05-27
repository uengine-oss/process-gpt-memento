"""workbook_card 추출 — dataset role 자료(엑셀/CSV) 업로드 시 메타 요약.

data-analyst 서브에이전트가 코드 실행 전 *어디에 무엇이 있는지* 파악할 수 있도록
시트 인벤토리·컬럼·dtype·기본 통계·샘플 5행·named ranges·수식 유무를 추출해 JSON 으로 반환.

원본 파일은 storage 그대로 보관 — 본 모듈은 *읽기 전용*. 분석 시점에 sandbox 가
원본을 다시 받아 쓴다.

설계 의도:
- LLM 컨텍스트에 들어가는 카드는 작아야 한다 → 시트당 ~수백 토큰 상한.
- 컬럼명·범주값은 *정확 표기 그대로* (LLM 환각 방지의 핵심).
- xlsx 의 *수식 보존* 정보(has_formulas, named_ranges)는 재무모델 점검 질문에 필수.
- 비정형 시트(병합셀·다중헤더)는 warnings 로 표시하고 강제 파싱 시도하지 않는다.
"""
from __future__ import annotations

import io
import logging
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

# 카드 크기 가드 — LLM 컨텍스트 보호용 상한
MAX_SAMPLE_ROWS = 5
MAX_TOP_CATEGORIES = 10
MAX_COLS_PER_SHEET = 50  # 그 이상은 잘라내고 warning
NUMERIC_DTYPES = ("int", "float")


def _coerce_value(v: Any) -> Any:
    """JSON 직렬화 가능한 형태로 값 정규화 (Timestamp, NaN, np.float 등)."""
    import math
    if v is None:
        return None
    try:
        import pandas as pd
        if isinstance(v, pd.Timestamp):
            return v.isoformat()
        if pd.isna(v):
            return None
    except Exception:
        pass
    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
        return None
    if isinstance(v, (int, float, str, bool)):
        return v
    return str(v)


def _summarize_dataframe(df: "Any", sheet_name: str) -> Dict[str, Any]:
    """단일 DataFrame → sheet card dict."""
    import pandas as pd  # 지연 import — 모듈 부팅 비용 회피

    n_rows = int(df.shape[0])
    n_cols = int(df.shape[1])

    cols_truncated = False
    if n_cols > MAX_COLS_PER_SHEET:
        df = df.iloc[:, :MAX_COLS_PER_SHEET]
        cols_truncated = True

    header: List[Dict[str, Any]] = []
    numeric_stats: Dict[str, Dict[str, Any]] = {}
    categorical_stats: Dict[str, Dict[str, Any]] = {}

    for col in df.columns:
        col_name = str(col)
        series = df[col]
        dtype = str(series.dtype)
        null_pct = float(series.isna().mean()) if n_rows > 0 else 0.0
        header.append({
            "name": col_name,
            "dtype": dtype,
            "null_pct": round(null_pct, 4),
        })

        if pd.api.types.is_numeric_dtype(series):
            non_null = series.dropna()
            if len(non_null) > 0:
                numeric_stats[col_name] = {
                    "min": _coerce_value(non_null.min()),
                    "max": _coerce_value(non_null.max()),
                    "mean": _coerce_value(round(float(non_null.mean()), 6)),
                }
        else:
            non_null = series.dropna()
            if len(non_null) > 0 and len(non_null) <= 100000:
                vc = non_null.astype(str).value_counts().head(MAX_TOP_CATEGORIES)
                categorical_stats[col_name] = {
                    "n_unique": int(non_null.nunique()),
                    "top": [
                        {"value": str(idx), "count": int(cnt)}
                        for idx, cnt in vc.items()
                    ],
                }

    sample = df.head(MAX_SAMPLE_ROWS).fillna("").astype(object)
    sample_rows: List[List[Any]] = [
        [_coerce_value(v) for v in row]
        for row in sample.values.tolist()
    ]

    card: Dict[str, Any] = {
        "name": sheet_name,
        "n_rows": n_rows,
        "n_cols": n_cols,
        "header": header,
        "sample_rows": sample_rows,
    }
    if numeric_stats:
        card["numeric_stats"] = numeric_stats
    if categorical_stats:
        card["categorical_stats"] = categorical_stats
    if cols_truncated:
        card["warnings"] = [f"컬럼 {n_cols}개 중 앞 {MAX_COLS_PER_SHEET}개만 카드에 수록"]
    return card


def extract_workbook_card(file_content: bytes, file_name: str) -> Dict[str, Any]:
    """xlsx/xls/csv/tsv → workbook_card dict.

    실패 시 ``{"kind": "unknown", "error": "..."}`` 반환 — 호출측은 인덱싱 실패로 처리하지 말고
    카드를 비워둔 채 indexed 로 마크하는 게 안전 (분석 시점에 sandbox 가 원본 재처리).
    """
    name_lower = (file_name or "").lower()
    warnings: List[str] = []

    try:
        if name_lower.endswith(".csv") or name_lower.endswith(".tsv"):
            return _extract_csv(file_content, file_name)

        if name_lower.endswith((".xlsx", ".xls")):
            return _extract_xlsx(file_content, file_name, warnings)

        return {"kind": "unknown", "error": f"지원하지 않는 확장자: {file_name}"}
    except Exception as e:
        logger.exception("[workbook_card] extraction failed for %s: %s", file_name, e)
        return {"kind": "unknown", "error": f"추출 실패: {e}"}


def _extract_csv(file_content: bytes, file_name: str) -> Dict[str, Any]:
    import pandas as pd

    sep = "\t" if file_name.lower().endswith(".tsv") else ","
    # 인코딩 자동 시도: utf-8 → cp949 → latin-1
    last_err: Optional[Exception] = None
    for enc in ("utf-8", "utf-8-sig", "cp949", "latin-1"):
        try:
            df = pd.read_csv(io.BytesIO(file_content), sep=sep, encoding=enc)
            sheet = _summarize_dataframe(df, sheet_name="(csv)")
            return {
                "kind": "csv",
                "encoding": enc,
                "sheets": [sheet],
            }
        except UnicodeDecodeError as e:
            last_err = e
            continue
        except Exception as e:
            last_err = e
            break
    return {"kind": "csv", "error": f"CSV 파싱 실패: {last_err}"}


def _extract_xlsx(file_content: bytes, file_name: str, warnings: List[str]) -> Dict[str, Any]:
    import pandas as pd
    from openpyxl import load_workbook

    bio = io.BytesIO(file_content)
    # data_only=False — 수식 그 자체를 보존 (수식 감지·named ranges 추출용)
    try:
        wb_meta = load_workbook(bio, data_only=False, read_only=False)
    except Exception as e:
        return {"kind": "workbook", "error": f"openpyxl 로드 실패: {e}"}

    named_ranges: List[str] = []
    try:
        named_ranges = sorted([str(n) for n in wb_meta.defined_names])
    except Exception:
        pass

    has_formulas_per_sheet: Dict[str, bool] = {}
    has_merged_per_sheet: Dict[str, bool] = {}
    for ws in wb_meta.worksheets:
        has_formula = False
        try:
            # 첫 100행 정도만 훑어서 수식 셀 존재 여부 확인 (전수 스캔 비용 방지)
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 1, 200), values_only=False):
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and val.startswith("="):
                        has_formula = True
                        break
                if has_formula:
                    break
        except Exception:
            pass
        has_formulas_per_sheet[ws.title] = has_formula
        try:
            has_merged_per_sheet[ws.title] = bool(list(ws.merged_cells.ranges))
        except Exception:
            has_merged_per_sheet[ws.title] = False

    # pandas 로 데이터 읽기 — 별도 stream (openpyxl 핸들과 분리)
    sheets_card: List[Dict[str, Any]] = []
    bio2 = io.BytesIO(file_content)
    try:
        all_sheets = pd.read_excel(bio2, sheet_name=None, engine="openpyxl")
    except Exception as e:
        warnings.append(f"pandas read_excel 실패 — 시트별 본문 요약 누락: {e}")
        all_sheets = {}

    for sheet_name in wb_meta.sheetnames:
        df = all_sheets.get(sheet_name)
        if df is None:
            sheets_card.append({
                "name": sheet_name,
                "n_rows": 0,
                "n_cols": 0,
                "header": [],
                "warnings": ["pandas 로 본문을 읽지 못함 (병합셀/다중헤더 가능)"],
                "has_formulas": has_formulas_per_sheet.get(sheet_name, False),
                "has_merged_cells": has_merged_per_sheet.get(sheet_name, False),
            })
            continue
        card = _summarize_dataframe(df, sheet_name)
        card["has_formulas"] = has_formulas_per_sheet.get(sheet_name, False)
        card["has_merged_cells"] = has_merged_per_sheet.get(sheet_name, False)
        if card["has_merged_cells"]:
            card.setdefault("warnings", []).append("병합셀 존재 — 분석 시 헤더 정규화 필요할 수 있음")
        sheets_card.append(card)

    try:
        wb_meta.close()
    except Exception:
        pass

    out: Dict[str, Any] = {
        "kind": "workbook",
        "sheets": sheets_card,
    }
    if named_ranges:
        out["named_ranges"] = named_ranges
    if warnings:
        out["warnings"] = warnings
    return out
